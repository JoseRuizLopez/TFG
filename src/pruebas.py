import re

from openpyxl import Workbook

from utils.utils_plot import generate_boxplot_from_csvs


def generate_excel_by_salidas(
    # Define la lista de términos a buscar y una lista para almacenar los resultados por algoritmo
    textos_a_buscar=["Accuracy", "Precision", "Recall", "F1-score", "Fitness check:"],
    resultados_algoritmos=[],

    # Variables para manejar el procesamiento
    resultados_generales={},
    fitness_checks_globales=[],  # Lista para almacenar todos los resultados de "Fitness Check" con nombre del algoritmo
    en_fitness_check=False,
    nombre_algoritmo=None,
):
    # Abre el archivo y procesa las líneas
    with open('results/salidas/salida_2024-11-02_23-07_task_0.txt', 'r') as file_read:
        for linea in file_read:
            # Detecta el inicio de un nuevo algoritmo
            if linea.startswith("--------"):
                # Guarda los resultados del algoritmo anterior, si existen
                if nombre_algoritmo:
                    resultados_algoritmos.append({
                        "algoritmo": nombre_algoritmo.replace('-', ''),
                        "generales": resultados_generales
                    })
                # Restablece las variables para el nuevo algoritmo
                nombre_algoritmo = linea.replace('-', '').strip()
                resultados_generales = {texto: [] for texto in textos_a_buscar if texto != "Fitness check:"}
                en_fitness_check = False
                continue

            # Verifica si estamos en un bloque de "Fitness check"
            if "Fitness check:" in linea:
                en_fitness_check = True
                continue

            # Procesa cada término en la línea según si estamos en "Fitness check" o no
            for texto in resultados_generales.keys():
                if texto in linea:
                    # Usa regex para extraer el número
                    match = re.search(r'{}:\s*([\d.]+)'.format(re.escape(texto)), linea)
                    if match:
                        valor = float(match.group(1))
                        # Guarda en el diccionario correspondiente
                        if en_fitness_check:
                            # Agrega el valor de Fitness Check junto con el nombre del algoritmo
                            fitness_checks_globales.append({
                                "algoritmo": nombre_algoritmo.replace('-', ''),
                                "metrica": texto,
                                "valor": valor
                            })
                        else:
                            resultados_generales[texto].append(valor)

    # Guarda el último algoritmo
    if nombre_algoritmo:
        resultados_algoritmos.append({
            "algoritmo": nombre_algoritmo.replace('-', ''),
            "generales": resultados_generales
        })

    # Crea el archivo Excel usando openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Algoritmos"

    # Define una posición inicial en la hoja de Excel
    columna_actual = 1
    max_filas_generales = 1

    # Agrega los datos de cada algoritmo en bloques de columnas
    for resultado in resultados_algoritmos:
        # Escribe el nombre del algoritmo en la primera fila del bloque
        ws.cell(row=1, column=columna_actual, value=resultado["algoritmo"])

        # Encabezado para métricas generales en la fila 2
        ws.cell(row=2, column=columna_actual, value="Accuracy")
        ws.cell(row=2, column=columna_actual + 1, value="Precision")
        ws.cell(row=2, column=columna_actual + 2, value="Recall")
        ws.cell(row=2, column=columna_actual + 3, value="F1-score")

        # Rellena los valores generales en las filas debajo del encabezado
        max_filas_generales = max(len(valores) for valores in resultado["generales"].values() if valores)
        for i in range(max_filas_generales):
            for j, metrica in enumerate(["Accuracy", "Precision", "Recall", "F1-score"]):
                valor = resultado["generales"].get(metrica, [])
                ws.cell(row=3 + i, column=columna_actual + j, value=valor[i] if i < len(valor) else None)

        # Mueve a la siguiente columna para el próximo algoritmo
        columna_actual += 6  # Espacio entre algoritmos para mejor separación

    # Agrega una sección para todos los resultados de Fitness Check al final de todos los algoritmos
    fila_fitness_global = 3 + max_filas_generales + 2  # Se asegura de estar después de los resultados generales

    # Encabezado de Fitness Check
    ws.cell(row=fila_fitness_global, column=1, value="Fitness Check Global")

    # Encabezado para métricas de Fitness Check incluyendo el nombre del algoritmo
    ws.cell(row=fila_fitness_global + 1, column=1, value="Algoritmo")
    ws.cell(row=fila_fitness_global + 1, column=2, value="Accuracy")
    ws.cell(row=fila_fitness_global + 1, column=3, value="Precision")
    ws.cell(row=fila_fitness_global + 1, column=4, value="Recall")
    ws.cell(row=fila_fitness_global + 1, column=5, value="F1-score")

    # Rellena los valores de Fitness Check globales en las filas debajo del encabezado
    fila_actual = fila_fitness_global + 2
    algoritmos_fitness = sorted(fitness_checks_globales, key=lambda x: x["algoritmo"])  # Ordena por algoritmo

    # Creamos una estructura para organizar las métricas por algoritmo
    fitness_organizado = {}
    for item in algoritmos_fitness:
        fitness_organizado.setdefault(item["algoritmo"], {}).setdefault(item["metrica"], []).append(item["valor"])

    for algoritmo, metricas in fitness_organizado.items():
        ws.cell(row=fila_actual, column=1, value=algoritmo)  # Nombre del algoritmo
        for j, metrica in enumerate(["Accuracy", "Precision", "Recall", "F1-score"], start=2):
            valores = metricas.get(metrica, [])
            for i, valor in enumerate(valores):
                ws.cell(row=fila_actual + i, column=j, value=valor)
        fila_actual += max(
            len(v) for v in metricas.values())  # Avanza tantas filas como valores haya para ese algoritmo

    # Guarda el archivo Excel
    wb.save('resultados_algoritmos_horizontal_con_fitness_unificado.xlsx')


if __name__ == "__main__":
    # generate_excel_by_salidas()

    read_csvs = [
        "results/csvs/ultima-prueba-PAINTING/2025-03-17_12-26/task_0.csv",
        "results/csvs/ultima-prueba-PAINTING/2025-03-17_12-26/task_1.csv",
        "results/csvs/ultima-prueba-PAINTING/2025-03-17_12-26/task_2.csv",
        "results/csvs/ultima-prueba-PAINTING/2025-03-17_12-26/task_3.csv",
        "results/csvs/ultima-prueba-PAINTING/2025-03-17_12-26/task_4.csv",
    ]
    generate_boxplot_from_csvs(read_csvs, None, modelo_name="mobilenet")
